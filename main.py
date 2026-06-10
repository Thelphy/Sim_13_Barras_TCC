import sys
import math
import json
from PyQt6.QtWidgets import QApplication, QMessageBox, QTableWidgetItem, QFileDialog
from openpyxl import Workbook, load_workbook
from PyQt6.QtCore import QThread, pyqtSignal, Qt, QSettings

from data_models import SystemState, BusData, LineData, CableConfig
from ui_main import MainWindowUI
from engine_sep import PowerSystemEngine
from plot_utils import populate_table

def safe_float(val_str):
    if isinstance(val_str, str):
        val_str = val_str.replace(',', '.')
    v = float(val_str)
    if not math.isfinite(v):
        raise ValueError(f"Value '{val_str}' is not a finite float.")
    return v

class SimulationThread(QThread):
    finished = pyqtSignal(object)

    def __init__(self, state, target_bus):
        super().__init__()
        self.state = state
        self.target_bus = target_bus

    def run(self):
        engine = PowerSystemEngine()
        engine.build_network(self.state)
        engine.run_power_flow()

        
        if engine.results.success:
            engine.run_modal_analysis()
        if self.target_bus:
            engine.generate_pv_curve(self.target_bus)
        self.finished.emit(engine.results)

class MainController:
    def __init__(self):
        self.app = QApplication(sys.argv)
        self.ui = MainWindowUI()
        self.state = SystemState()
        self.scenarios = {}

        self.init_default_data()
        self.load_settings()
        self.load_scenarios()
        self.setup_connections()

        # Initial draw
        self.update_diagram()
        self.populate_params_tables()
        self.populate_target_combo()
        self.update_scenarios_combo()


    def save_settings(self):
        settings = QSettings("SimuladorSEP", "Parametros")

        # Save buses
        buses_data = {}
        for bus_id, bus in self.state.buses.items():
            buses_data[bus_id] = {
                "p_load_kw": bus.p_load_kw,
                "q_load_kvar": bus.q_load_kvar,
                "p_gen_kw": bus.p_gen_kw
            }
        settings.setValue("buses", json.dumps(buses_data))

        # Save lines
        lines_data = {}
        for line_id, line in self.state.lines.items():
            lines_data[line_id] = {
                "r_ohm_per_km": line.r_ohm_per_km,
                "x_ohm_per_km": line.x_ohm_per_km,
                "length_km": line.length_km,
                "sn_mva": line.sn_mva,
                "vk_percent": line.vk_percent,
                "vkr_percent": line.vkr_percent,
                "pfe_kw": line.pfe_kw,
                "i0_percent": line.i0_percent
            }
        settings.setValue("lines", json.dumps(lines_data))

        # Save cables
        cables_data = {}
        for c_name, cable in self.state.cables.items():
            cables_data[c_name] = {
                "r_ohm_per_km": cable.r_ohm_per_km,
                "x_ohm_per_km": cable.x_ohm_per_km
            }
        settings.setValue("cables", json.dumps(cables_data))

    def load_settings(self):
        settings = QSettings("SimuladorSEP", "Parametros")

        buses_str = settings.value("buses", "")
        if buses_str:
            try:
                buses_data = json.loads(buses_str)
                for bus_id_str, data in buses_data.items():
                    bus_id = int(bus_id_str)
                    if bus_id in self.state.buses:
                        self.state.buses[bus_id].p_load_kw = float(data.get("p_load_kw", self.state.buses[bus_id].p_load_kw))
                        self.state.buses[bus_id].q_load_kvar = float(data.get("q_load_kvar", self.state.buses[bus_id].q_load_kvar))
                        self.state.buses[bus_id].p_gen_kw = float(data.get("p_gen_kw", self.state.buses[bus_id].p_gen_kw))
            except Exception as e:
                print("Erro ao carregar parâmetros de barras:", e)

        lines_str = settings.value("lines", "")
        if lines_str:
            try:
                lines_data = json.loads(lines_str)
                for line_id_str, data in lines_data.items():
                    line_id = int(line_id_str)
                    if line_id in self.state.lines:
                        self.state.lines[line_id].r_ohm_per_km = float(data.get("r_ohm_per_km", self.state.lines[line_id].r_ohm_per_km))
                        self.state.lines[line_id].x_ohm_per_km = float(data.get("x_ohm_per_km", self.state.lines[line_id].x_ohm_per_km))
                        self.state.lines[line_id].length_km = float(data.get("length_km", self.state.lines[line_id].length_km))
                        self.state.lines[line_id].sn_mva = float(data.get("sn_mva", self.state.lines[line_id].sn_mva))
                        self.state.lines[line_id].vk_percent = float(data.get("vk_percent", self.state.lines[line_id].vk_percent))
                        self.state.lines[line_id].vkr_percent = float(data.get("vkr_percent", self.state.lines[line_id].vkr_percent))
                        self.state.lines[line_id].pfe_kw = float(data.get("pfe_kw", self.state.lines[line_id].pfe_kw))
                        self.state.lines[line_id].i0_percent = float(data.get("i0_percent", self.state.lines[line_id].i0_percent))
            except Exception as e:
                print("Erro ao carregar parâmetros de linhas:", e)
                
        cables_str = settings.value("cables", "")
        if cables_str:
            try:
                cables_data = json.loads(cables_str)
                for c_name, data in cables_data.items():
                    self.state.cables[c_name] = CableConfig(
                        name=c_name,
                        r_ohm_per_km=float(data.get("r_ohm_per_km", 0.1)),
                        x_ohm_per_km=float(data.get("x_ohm_per_km", 0.1))
                    )
            except Exception as e:
                print("Erro ao carregar parâmetros de cabos:", e)

    def load_scenarios(self):
        settings = QSettings("SimuladorSEP", "Cenarios")
        data = settings.value("data", "{}")
        try:
            self.scenarios = json.loads(data)
        except:
            self.scenarios = {}

    def save_scenarios(self):
        settings = QSettings("SimuladorSEP", "Cenarios")
        settings.setValue("data", json.dumps(self.scenarios))
        self.update_scenarios_combo()

    def update_scenarios_combo(self):
        self.ui.combo_scenarios.blockSignals(True)
        self.ui.combo_scenarios.clear()
        self.ui.combo_scenarios.addItem("-- Selecionar Cenário --")
        self.ui.combo_scenarios.addItems(list(self.scenarios.keys()))
        self.ui.combo_scenarios.blockSignals(False)

    def on_scenario_selected(self, idx):
        if idx <= 0:
            return
        name = self.ui.combo_scenarios.currentText()
        if name in self.scenarios:
            scen_data = self.scenarios[name]
            self.apply_scenario_data(scen_data)
            self.populate_params_tables()
            self.update_diagram()
            self.save_settings()
            self.ui.toast.show_toast(f"Cenário '{name}' carregado!", True, self.ui)

    def apply_scenario_data(self, scen_data):
        buses_data = scen_data.get("buses", {})
        for bus_id_str, data in buses_data.items():
            bus_id = int(bus_id_str)
            if bus_id in self.state.buses:
                self.state.buses[bus_id].p_load_kw = float(data.get("p_load_kw", self.state.buses[bus_id].p_load_kw))
                self.state.buses[bus_id].q_load_kvar = float(data.get("q_load_kvar", self.state.buses[bus_id].q_load_kvar))
                self.state.buses[bus_id].p_gen_kw = float(data.get("p_gen_kw", self.state.buses[bus_id].p_gen_kw))

        lines_data = scen_data.get("lines", {})
        for line_id_str, data in lines_data.items():
            line_id = int(line_id_str)
            if line_id in self.state.lines:
                self.state.lines[line_id].r_ohm_per_km = float(data.get("r_ohm_per_km", self.state.lines[line_id].r_ohm_per_km))
                self.state.lines[line_id].x_ohm_per_km = float(data.get("x_ohm_per_km", self.state.lines[line_id].x_ohm_per_km))
                self.state.lines[line_id].length_km = float(data.get("length_km", self.state.lines[line_id].length_km))
                self.state.lines[line_id].sn_mva = float(data.get("sn_mva", self.state.lines[line_id].sn_mva))
                self.state.lines[line_id].vk_percent = float(data.get("vk_percent", self.state.lines[line_id].vk_percent))
                self.state.lines[line_id].vkr_percent = float(data.get("vkr_percent", self.state.lines[line_id].vkr_percent))
                self.state.lines[line_id].pfe_kw = float(data.get("pfe_kw", self.state.lines[line_id].pfe_kw))
                self.state.lines[line_id].i0_percent = float(data.get("i0_percent", self.state.lines[line_id].i0_percent))

    def save_scenario_action(self):
        from PyQt6.QtWidgets import QInputDialog
        name, ok = QInputDialog.getText(self.ui, "Salvar Cenário", "Nome do Cenário:")
        if ok and name.strip():
            name = name.strip()
            # Serialize current parameters
            buses_data = {}
            for bus_id, bus in self.state.buses.items():
                buses_data[bus_id] = {
                    "p_load_kw": bus.p_load_kw,
                    "q_load_kvar": bus.q_load_kvar,
                    "p_gen_kw": bus.p_gen_kw
                }
            lines_data = {}
            for line_id, line in self.state.lines.items():
                lines_data[line_id] = {
                    "r_ohm_per_km": line.r_ohm_per_km,
                    "x_ohm_per_km": line.x_ohm_per_km,
                    "length_km": line.length_km,
                    "sn_mva": line.sn_mva,
                    "vk_percent": line.vk_percent,
                    "vkr_percent": line.vkr_percent,
                    "pfe_kw": line.pfe_kw,
                    "i0_percent": line.i0_percent
                }
            self.scenarios[name] = {"buses": buses_data, "lines": lines_data}
            self.save_scenarios()
            self.ui.toast.show_toast(f"Cenário '{name}' salvo!", True, self.ui)

    def delete_scenario_action(self):
        from PyQt6.QtWidgets import QInputDialog
        if not self.scenarios:
            QMessageBox.information(self.ui, "Excluir Cenário", "Nenhum cenário salvo.")
            return

        items = list(self.scenarios.keys())
        item, ok = QInputDialog.getItem(self.ui, "Excluir Cenário", "Selecione o cenário a excluir:", items, 0, False)
        if ok and item:
            del self.scenarios[item]
            self.save_scenarios()
            self.ui.toast.show_toast(f"Cenário '{item}' excluído!", True, self.ui)

    def init_default_data(self):
        # 13 Bus system based on IEEE 13 Node Test Feeder layout
        self.state.buses[650] = BusData(id=650, name="650 (Slack)", vn_kv=13.8, type='slack', v_target_pu=1.0)
        self.state.buses[632] = BusData(id=632, name="632", vn_kv=13.8, type='pq', p_load_kw=0, q_load_kvar=0)
        self.state.buses[645] = BusData(id=645, name="645", vn_kv=13.8, type='pq', p_load_kw=170, q_load_kvar=125)
        self.state.buses[646] = BusData(id=646, name="646", vn_kv=13.8, type='pq', p_load_kw=230, q_load_kvar=132)
        self.state.buses[633] = BusData(id=633, name="633", vn_kv=13.8, type='pq', p_load_kw=0, q_load_kvar=0)
        self.state.buses[634] = BusData(id=634, name="634", vn_kv=0.22, type='pq', p_load_kw=340, q_load_kvar=120)
        self.state.buses[671] = BusData(id=671, name="671", vn_kv=13.8, type='pq', p_load_kw=1155, q_load_kvar=660)
        self.state.buses[684] = BusData(id=684, name="684", vn_kv=13.8, type='pq', p_load_kw=0, q_load_kvar=0)
        self.state.buses[611] = BusData(id=611, name="611", vn_kv=13.8, type='pq', p_load_kw=170, q_load_kvar=-220)
        self.state.buses[652] = BusData(id=652, name="652", vn_kv=13.8, type='pq', p_load_kw=128, q_load_kvar=86)
        self.state.buses[692] = BusData(id=692, name="692", vn_kv=13.8, type='pq', p_load_kw=170, q_load_kvar=151)
        self.state.buses[675] = BusData(id=675, name="675", vn_kv=13.8, type='pq', p_load_kw=843, q_load_kvar=-138)
        self.state.buses[680] = BusData(id=680, name="680", vn_kv=13.8, type='pq', p_load_kw=0, q_load_kvar=0)

        # Connect the buses
        # (id, from, to, length, r, x)
        self.state.lines[1] = LineData(1, 650, 632, 0.6096, 0.1155, 0.371, is_transformer=False)
        self.state.lines[2] = LineData(2, 632, 645, 0.1524, 0.3679, 0.4726)
        self.state.lines[3] = LineData(3, 645, 646, 0.0914, 0.3679, 0.4726)
        self.state.lines[4] = LineData(4, 632, 633, 0.1524, 0.3679, 0.4726)
        self.state.lines[5] = LineData(5, 633, 634, 0.0, 0.0, 0.0, is_transformer=True, sn_mva=0.5, vk_percent=4.0, vkr_percent=1.0, pfe_kw=0.0, i0_percent=0.5)
        self.state.lines[6] = LineData(6, 632, 671, 0.6096, 0.1155, 0.371)
        self.state.lines[7] = LineData(7, 671, 684, 0.0914, 0.3679, 0.4726)
        self.state.lines[8] = LineData(8, 684, 611, 0.0914, 0.3679, 0.4726)
        self.state.lines[9] = LineData(9, 684, 652, 0.2438, 0.3679, 0.4726)
        self.state.lines[10] = LineData(10, 671, 692, 0.01, 0.01, 0.01) # switch (avoid divide by zero and ill-conditioning)
        self.state.lines[11] = LineData(11, 692, 675, 0.1524, 0.3679, 0.4726)
        self.state.lines[12] = LineData(12, 671, 680, 0.3048, 0.1155, 0.371)

        self.populate_params_tables()

    def setup_connections(self):
        self.ui.btn_simulate.clicked.connect(self.run_simulation)
        self.ui.diagram_view.data_updated.connect(self.on_diagram_data_updated)
        self.ui.btn_save_params.clicked.connect(self.save_params)
        self.ui.btn_export_params.clicked.connect(self.export_params)
        self.ui.btn_import_params.clicked.connect(self.import_params)
        self.ui.btn_export_results.clicked.connect(self.export_results)
        self.ui.btn_export_plot.clicked.connect(self.export_plot)
        self.ui.app_closed.connect(self.save_settings)
        self.ui.btn_save_scenario.clicked.connect(self.save_scenario_action)
        self.ui.btn_delete_scenario.clicked.connect(self.delete_scenario_action)
        self.ui.combo_scenarios.currentIndexChanged.connect(self.on_scenario_selected)
        
        self.ui.btn_add_cable.clicked.connect(self.add_cable)
        self.ui.btn_remove_cable.clicked.connect(self.remove_cable)

    def on_diagram_data_updated(self):
        self.update_diagram()
        self.populate_params_tables()
        self.save_settings()
        
    def add_cable(self):
        from PyQt6.QtWidgets import QInputDialog
        name, ok = QInputDialog.getText(self.ui, "Novo Cabo", "Nome do Cabo:")
        if ok and name and name not in self.state.cables:
            self.state.cables[name] = CableConfig(name=name, r_ohm_per_km=0.1, x_ohm_per_km=0.1)
            self.populate_params_tables()
            self.save_settings()
            
    def remove_cable(self):
        from PyQt6.QtWidgets import QInputDialog
        if not self.state.cables:
            return
        items = list(self.state.cables.keys())
        item, ok = QInputDialog.getItem(self.ui, "Remover Cabo", "Selecione o cabo:", items, 0, False)
        if ok and item in self.state.cables:
            del self.state.cables[item]
            self.populate_params_tables()
            self.save_settings()

    def populate_params_tables(self):
        from PyQt6.QtWidgets import QTableWidgetItem
        # Populate Cables
        self.ui.table_cables.setRowCount(len(self.state.cables))
        self.ui.table_cables.setColumnCount(3)
        self.ui.table_cables.setHorizontalHeaderLabels(["Nome", "R (ohm/km)", "X (ohm/km)"])
        for i, (name, cable) in enumerate(self.state.cables.items()):
            self.ui.table_cables.setItem(i, 0, QTableWidgetItem(name))
            self.ui.table_cables.setItem(i, 1, QTableWidgetItem(str(cable.r_ohm_per_km)))
            self.ui.table_cables.setItem(i, 2, QTableWidgetItem(str(cable.x_ohm_per_km)))

        # Populate Buses
        self.ui.table_params_buses.setRowCount(len(self.state.buses))
        self.ui.table_params_buses.setColumnCount(4)
        self.ui.table_params_buses.setHorizontalHeaderLabels(["ID", "P Load (kW)", "Q Load (kVAr)", "Geração (kW)"])
        for i, (bus_id, bus) in enumerate(self.state.buses.items()):
            self.ui.table_params_buses.setItem(i, 0, QTableWidgetItem(str(bus.id)))
            self.ui.table_params_buses.setItem(i, 1, QTableWidgetItem(str(bus.p_load_kw)))
            self.ui.table_params_buses.setItem(i, 2, QTableWidgetItem(str(bus.q_load_kvar)))
            self.ui.table_params_buses.setItem(i, 3, QTableWidgetItem(str(bus.p_gen_kw)))

        # Populate Lines
        from PyQt6.QtWidgets import QComboBox
        normal_lines = [l for l in self.state.lines.values() if not l.is_transformer]
        self.ui.table_params_lines.setRowCount(len(normal_lines))
        self.ui.table_params_lines.setColumnCount(5)
        self.ui.table_params_lines.setHorizontalHeaderLabels(["ID", "R (ohm/km)", "X (ohm/km)", "Length (km)", "Cabo"])
        for i, line in enumerate(normal_lines):
            item = QTableWidgetItem(f"{line.from_bus} - {line.to_bus}")
            item.setData(Qt.ItemDataRole.UserRole, line.id)
            self.ui.table_params_lines.setItem(i, 0, item)
            self.ui.table_params_lines.setItem(i, 1, QTableWidgetItem(str(line.r_ohm_per_km)))
            self.ui.table_params_lines.setItem(i, 2, QTableWidgetItem(str(line.x_ohm_per_km)))
            self.ui.table_params_lines.setItem(i, 3, QTableWidgetItem(str(line.length_km)))
            
            combo = QComboBox()
            combo.addItem("Personalizado")
            for c_name in self.state.cables.keys():
                combo.addItem(c_name)
            
            # Select correct cable if matches
            for c_name, c_data in self.state.cables.items():
                if abs(c_data.r_ohm_per_km - line.r_ohm_per_km) < 1e-4 and abs(c_data.x_ohm_per_km - line.x_ohm_per_km) < 1e-4:
                    combo.setCurrentText(c_name)
                    break
            
            # Connect
            def on_cable_selected(text, row=i):
                if text in self.state.cables:
                    c = self.state.cables[text]
                    self.ui.table_params_lines.item(row, 1).setText(str(c.r_ohm_per_km))
                    self.ui.table_params_lines.item(row, 2).setText(str(c.x_ohm_per_km))
            
            combo.currentTextChanged.connect(on_cable_selected)
            self.ui.table_params_lines.setCellWidget(i, 4, combo)

        # Populate Transformers
        trafos = [l for l in self.state.lines.values() if l.is_transformer]
        self.ui.table_params_trafos.setRowCount(len(trafos))
        self.ui.table_params_trafos.setColumnCount(6)
        self.ui.table_params_trafos.setHorizontalHeaderLabels(["ID", "Sn (MVA)", "vk (%)", "vkr (%)", "pfe (kW)", "i0 (%)"])
        for i, trafo in enumerate(trafos):
            item = QTableWidgetItem(f"{trafo.from_bus} - {trafo.to_bus}")
            item.setData(Qt.ItemDataRole.UserRole, trafo.id)
            self.ui.table_params_trafos.setItem(i, 0, item)
            self.ui.table_params_trafos.setItem(i, 1, QTableWidgetItem(str(trafo.sn_mva)))
            self.ui.table_params_trafos.setItem(i, 2, QTableWidgetItem(str(trafo.vk_percent)))
            self.ui.table_params_trafos.setItem(i, 3, QTableWidgetItem(str(trafo.vkr_percent)))
            self.ui.table_params_trafos.setItem(i, 4, QTableWidgetItem(str(trafo.pfe_kw)))
            self.ui.table_params_trafos.setItem(i, 5, QTableWidgetItem(str(trafo.i0_percent)))
            
        self.ui.table_cables.resizeColumnsToContents()
        self.ui.table_params_buses.resizeColumnsToContents()
        self.ui.table_params_lines.resizeColumnsToContents()
        self.ui.table_params_trafos.resizeColumnsToContents()
        
        self.adjust_table_size(self.ui.table_cables)
        self.adjust_table_size(self.ui.table_params_buses)
        self.adjust_table_size(self.ui.table_params_lines)
        self.adjust_table_size(self.ui.table_params_trafos)

    def adjust_table_size(self, table):
        h = table.horizontalHeader().height()
        for i in range(table.rowCount()):
            h += table.rowHeight(i)
        table.setMinimumHeight(h + 10)
        
        w = table.verticalHeader().width() if table.verticalHeader().isVisible() else 0
        for i in range(table.columnCount()):
            w += table.columnWidth(i)
        table.setMinimumWidth(w + 30)

    def save_params(self):
        try:
            for i in range(self.ui.table_cables.rowCount()):
                name = self.ui.table_cables.item(i, 0).text()
                if name in self.state.cables:
                    self.state.cables[name].r_ohm_per_km = safe_float(self.ui.table_cables.item(i, 1).text())
                    self.state.cables[name].x_ohm_per_km = safe_float(self.ui.table_cables.item(i, 2).text())
                    
            for i in range(self.ui.table_params_buses.rowCount()):
                bus_id = int(self.ui.table_params_buses.item(i, 0).text())
                if bus_id in self.state.buses:
                    self.state.buses[bus_id].p_load_kw = safe_float(self.ui.table_params_buses.item(i, 1).text())
                    self.state.buses[bus_id].q_load_kvar = safe_float(self.ui.table_params_buses.item(i, 2).text())
                    self.state.buses[bus_id].p_gen_kw = safe_float(self.ui.table_params_buses.item(i, 3).text())

            for i in range(self.ui.table_params_lines.rowCount()):
                line_id = self.ui.table_params_lines.item(i, 0).data(Qt.ItemDataRole.UserRole)
                if line_id in self.state.lines:
                    self.state.lines[line_id].r_ohm_per_km = safe_float(self.ui.table_params_lines.item(i, 1).text())
                    self.state.lines[line_id].x_ohm_per_km = safe_float(self.ui.table_params_lines.item(i, 2).text())
                    self.state.lines[line_id].length_km = safe_float(self.ui.table_params_lines.item(i, 3).text())

            for i in range(self.ui.table_params_trafos.rowCount()):
                trafo_id = self.ui.table_params_trafos.item(i, 0).data(Qt.ItemDataRole.UserRole)
                if trafo_id in self.state.lines:
                    self.state.lines[trafo_id].sn_mva = safe_float(self.ui.table_params_trafos.item(i, 1).text())
                    self.state.lines[trafo_id].vk_percent = safe_float(self.ui.table_params_trafos.item(i, 2).text())
                    self.state.lines[trafo_id].vkr_percent = safe_float(self.ui.table_params_trafos.item(i, 3).text())
                    self.state.lines[trafo_id].pfe_kw = safe_float(self.ui.table_params_trafos.item(i, 4).text())
                    self.state.lines[trafo_id].i0_percent = safe_float(self.ui.table_params_trafos.item(i, 5).text())

            self.update_diagram()
            self.save_settings()
            QMessageBox.information(self.ui, "Sucesso", "Parâmetros salvos com sucesso!")
        except ValueError:
            QMessageBox.warning(self.ui, "Erro", "Valores inválidos inseridos. Use apenas números.")

    def export_params(self):
        import datetime
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        default_name = f"parametros_{timestamp}.xlsx"
        filename, _ = QFileDialog.getSaveFileName(self.ui, "Exportar Parâmetros", default_name, "Excel Files (*.xlsx)")
        if not filename:
            return

        try:
            wb = Workbook()
            ws = wb.active

            ws.append(["[Buses]"])
            ws.append(["ID", "P Load (kW)", "Q Load (kVAr)", "Geração (kW)"])
            for bus_id, bus in self.state.buses.items():
                ws.append([bus_id, bus.p_load_kw, bus.q_load_kvar, bus.p_gen_kw])

            ws.append([])
            ws.append(["[Lines]"])
            ws.append(["ID", "R (ohm/km)", "X (ohm/km)", "Length (km)"])
            for line_id, line in self.state.lines.items():
                if not line.is_transformer:
                    ws.append([line.id, line.r_ohm_per_km, line.x_ohm_per_km, line.length_km])

            ws.append([])
            ws.append(["[Transformers]"])
            ws.append(["ID", "Sn (MVA)", "vk (%)", "vkr (%)", "pfe (kW)", "i0 (%)"])
            for line_id, line in self.state.lines.items():
                if line.is_transformer:
                    ws.append([line.id, line.sn_mva, line.vk_percent, line.vkr_percent, line.pfe_kw, line.i0_percent])

            wb.save(filename)
            QMessageBox.information(self.ui, "Sucesso", "Parâmetros exportados com sucesso!")
        except Exception as e:
            QMessageBox.critical(self.ui, "Erro", f"Erro ao exportar: {str(e)}")

    def import_params(self):
        filename, _ = QFileDialog.getOpenFileName(self.ui, "Importar Parâmetros", "", "Excel Files (*.xlsx)")
        if not filename:
            return

        try:
            wb = load_workbook(filename)
            ws = wb.active
            mode = None

            for row in ws.iter_rows(values_only=True):
                if not row or row[0] is None:
                    continue
                if row[0] == "[Buses]":
                    mode = "buses"
                    continue
                elif row[0] == "[Lines]":
                    mode = "lines"
                    continue
                elif row[0] == "[Transformers]":
                    mode = "transformers"
                    continue

                if row[0] == "ID":
                    continue

                if mode == "buses":
                    bus_id = int(row[0])
                    if bus_id in self.state.buses:
                        self.state.buses[bus_id].p_load_kw = safe_float(row[1])
                        self.state.buses[bus_id].q_load_kvar = safe_float(row[2])
                        self.state.buses[bus_id].p_gen_kw = safe_float(row[3])
                elif mode == "lines":
                    line_id = int(row[0])
                    if line_id in self.state.lines:
                        self.state.lines[line_id].r_ohm_per_km = safe_float(row[1])
                        self.state.lines[line_id].x_ohm_per_km = safe_float(row[2])
                        self.state.lines[line_id].length_km = safe_float(row[3])
                elif mode == "transformers":
                    trafo_id = int(row[0])
                    if trafo_id in self.state.lines:
                        self.state.lines[trafo_id].sn_mva = safe_float(row[1])
                        self.state.lines[trafo_id].vk_percent = safe_float(row[2])
                        self.state.lines[trafo_id].vkr_percent = safe_float(row[3])
                        self.state.lines[trafo_id].pfe_kw = safe_float(row[4])
                        self.state.lines[trafo_id].i0_percent = safe_float(row[5])

            self.populate_params_tables()
            self.update_diagram()
            self.save_settings()
            QMessageBox.information(self.ui, "Sucesso", "Parâmetros importados com sucesso!")
        except Exception as e:
            QMessageBox.critical(self.ui, "Erro", f"Erro ao importar: {str(e)}")



    def export_plot(self):
        import datetime
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        default_name = f"curva_pv_{timestamp}.png"
        filename, _ = QFileDialog.getSaveFileName(self.ui, "Exportar Gráfico", default_name, "Images (*.png *.jpg *.jpeg)")
        if not filename:
            return

        try:
            self.ui.pv_plot.export_plot(filename)
            QMessageBox.information(self.ui, "Sucesso", "Gráfico exportado com sucesso!")
        except Exception as e:
            QMessageBox.critical(self.ui, "Erro", f"Erro ao exportar: {str(e)}")

    def export_results(self):
        import datetime
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        default_name = f"resultados_{timestamp}.xlsx"
        filename, _ = QFileDialog.getSaveFileName(self.ui, "Exportar Resultados", default_name, "Excel Files (*.xlsx)")
        if not filename:
            return

        try:
            wb = Workbook()
            ws = wb.active

            # Export Bus Results
            ws.append(["[Fluxo de Potência (Tensões Nodais)]"])
            headers = []
            for j in range(self.ui.table_bus_results.columnCount()):
                headers.append(self.ui.table_bus_results.horizontalHeaderItem(j).text())
            ws.append(headers)

            for i in range(self.ui.table_bus_results.rowCount()):
                row_data = []
                for j in range(self.ui.table_bus_results.columnCount()):
                    item = self.ui.table_bus_results.item(i, j)
                    row_data.append(item.text() if item else "")
                ws.append(row_data)
            ws.append([])

            # Export Line Results
            ws.append(["[Fluxo nas Linhas]"])
            headers = []
            for j in range(self.ui.table_line_results.columnCount()):
                headers.append(self.ui.table_line_results.horizontalHeaderItem(j).text())
            ws.append(headers)

            for i in range(self.ui.table_line_results.rowCount()):
                row_data = []
                for j in range(self.ui.table_line_results.columnCount()):
                    item = self.ui.table_line_results.item(i, j)
                    row_data.append(item.text() if item else "")
                ws.append(row_data)
            ws.append([])

            # Export Modal Analysis Results
            ws.append(["[Análise Modal]"])
            headers = []
            for j in range(self.ui.table_modal_results.columnCount()):
                headers.append(self.ui.table_modal_results.horizontalHeaderItem(j).text())
            ws.append(headers)

            for i in range(self.ui.table_modal_results.rowCount()):
                row_data = []
                for j in range(self.ui.table_modal_results.columnCount()):
                    item = self.ui.table_modal_results.item(i, j)
                    row_data.append(item.text() if item else "")
                ws.append(row_data)

            wb.save(filename)
            QMessageBox.information(self.ui, "Sucesso", "Resultados exportados com sucesso!")
        except Exception as e:
            QMessageBox.critical(self.ui, "Erro", f"Erro ao exportar: {str(e)}")

    def populate_target_combo(self):
        self.ui.combo_target_bus.clear()
        for bus in self.state.buses.values():
            if bus.type == 'pq':
                self.ui.combo_target_bus.addItem(bus.name)

    def update_diagram(self):
        self.ui.diagram_view.draw_network(self.state)

    def run_simulation(self):
        self.ui.btn_simulate.setEnabled(False)
        self.ui.btn_simulate.setText("Simulando...")
        self.ui.btn_simulate.setStyleSheet("background-color: #555555; color: #aaaaaa;")
        self.ui.progress_bar.setVisible(True)

        target_bus = self.ui.combo_target_bus.currentText()

        self.thread = SimulationThread(self.state, target_bus)
        self.thread.finished.connect(self.on_simulation_finished)
        self.thread.start()

    def on_simulation_finished(self, results):
        self.ui.btn_simulate.setEnabled(True)
        self.ui.btn_simulate.setText("Iniciar Simulação")
        self.ui.btn_simulate.setStyleSheet("")
        self.ui.progress_bar.setVisible(False)

        if not results.success:
            self.ui.toast.show_toast("Erro: A simulação falhou.", False, self.ui)
            return

        self.ui.toast.show_toast("Simulação concluída com sucesso!", True, self.ui)

        # Update UI Tables
        bus_headers = ["Barra", "V (PU)", "Ângulo (°)", "P (MW)", "Q (MVAr)"]
        populate_table(self.ui.table_bus_results, results.bus_results, bus_headers)

        line_headers = ["Linha", "P_in (MW)", "Q_in (MVAr)", "P_out (MW)", "Q_out (MVAr)", "Perda (MW)", "Carga (%)"]
        populate_table(self.ui.table_line_results, results.line_results, line_headers)

        modal_data = []
        if results.participation_factors:
            for i, (bus_name, pf) in enumerate(results.participation_factors.items()):
                eig_val = results.eigenvalues[i] if i < len(results.eigenvalues) else "-"
                modal_data.append([bus_name, f"{eig_val}", f"{pf:.4f}"])

        modal_headers = ["Barra", "Autovalor (Real)", "Fator de Participação"]
        populate_table(self.ui.table_modal_results, modal_data, modal_headers)
        
        self.adjust_table_size(self.ui.table_bus_results)
        self.adjust_table_size(self.ui.table_line_results)
        self.adjust_table_size(self.ui.table_modal_results)

        # Update Plot
        target_bus = self.ui.combo_target_bus.currentText()
        self.ui.pv_plot.plot_curve(results.pv_curve_p, results.pv_curve_v, target_bus)

    def run(self):
        self.ui.show()
        sys.exit(self.app.exec())

if __name__ == '__main__':
    controller = MainController()
    controller.run()
